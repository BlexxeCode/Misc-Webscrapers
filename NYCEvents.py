import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl import Workbook
import pandas as pd
from datetime import date
import time

browser = 'C://Users//andki//AppData//Local//Vivaldi//Application//vivaldi.exe'


path = 'C://Users//andki//OneDrive//Documents//Events.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

options = webdriver.ChromeOptions()
options.binary_location = browser

driver = webdriver.Chrome('C://Users//andki//Downloads//chromedriver//chromedriver.exe', options=options)


driver.get('https://www.nyc.gov/events/events-filter.html#page-1')
time.sleep(5)
num = 2

while True :
    events = driver.find_elements('css selector', 'li.event-item.span12')
    for event in events:
        title = event.find_element('xpath', './/h4').text
        date = event.find_element('css selector', 'span.date').text
        location = event.find_element('css selector','span.address').text
        details = event.find_element('xpath', ".//span[contains(@id, 'more')]").text
        print(title)
        print(date)
        print(location)
        print(details)

        ws['A' + str(num)].value = title
        ws['B' + str(num)].value = date
        ws['C' + str(num)].value = location
        ws['D' + str(num)].value = details
        wb.save(path)
        num += 1

    driver.find_element('css selector', 'a.page-link.next').click()

    time.sleep(3)



