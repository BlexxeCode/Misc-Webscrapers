import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl import Workbook
import pandas as pd
from datetime import date
import time

browser = 'C://Users//andki//AppData//Local//Vivaldi//Application//vivaldi.exe'


path = 'C://Users//andki//OneDrive//Documents//EGT_LinkedIn.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

options = webdriver.ChromeOptions()
options.binary_location = browser


url = input("LinkedIn URL")
Max = int(input('Amount of locations to obtain info from?'))
start = int(input("Row to start at?"))
num=start
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

driver.get(url)
time.sleep(5)
driver.find_element('css selector','a.main__sign-in-link').click()
time.sleep(3)
username = driver.find_element('id','username')
username.send_keys('andrewsaintvil6@gmail.com')

password = driver.find_element('id','password')
password.send_keys('7bG5rjYe8zN$')

time.sleep(2)
driver.find_element('css selector','button.btn__primary--large.from__button--floating').click()
time.sleep(20)


p = 1
while p <= Max:
    list1 = driver.find_elements('css selector', 'div.entity-result__content.entity-result__divider.pt3.pb3.t-12.t-black--light')
    for names in list1:
        website = names.find_element('css selector', 'a.app-aware-link').get_attribute('href')
        name = names.find_element('css selector', 'a.app-aware-link').text
        location = names.find_element('css selector', 'div.entity-result__secondary-subtitle.t-14.t-normal').text
        occupation = names.find_element('css selector', 'div.entity-result__primary-subtitle.t-14.t-black.t-normal').text
        experience = names.find_elements('css selector', 'div.linked-area.flex-1.cursor-pointer')
        experience[0].location_once_scrolled_into_view




        ws['A' + str(num)].value = str(date.today())
        ws['B' + str(num)].value = name
        ws['C' + str(num)].value = occupation
        ws['D' + str(num)].value = location
        try:
            exp = experience[1].text
            ws['E' + str(num)].value = experience[1].text
        except IndexError as e:
            exp = 'No summary'
            ws['E' + str(num)].value = exp
        ws['F' + str(num)].value = website

        wb.save(path)

        print(website)
        print(name)
        print(occupation)
        print(location)
        print(exp)

        num +=1
        p +=1

    time.sleep(5)
    if p >= Max:
        print("Done")
    #button = driver.find_element('css selector' ,"span.artdeco-button__text")
    #driver.execute_script("arguments[0].click();", button)
    time.sleep(3)



